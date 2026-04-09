import cv2
import numpy as np
import os
from datetime import datetime
from openpyxl import Workbook, load_workbook

# Configuration
size = 2
fn_dir = 'att_faces'  # Training dataset folder
(im_width, im_height) = (112, 92)
confidence_threshold = 500

# Excel file path
excel_file = "attendance.xlsx"

print('Training...')

# Create data holders
(images, labels, names, id) = ([], [], {}, 0)

# Load training images from subfolders
for (subdirs, dirs, files) in os.walk(fn_dir):
    for subdir in dirs:
        names[id] = subdir
        subject_path = os.path.join(fn_dir, subdir)
        for filename in os.listdir(subject_path):
            f_name, f_extension = os.path.splitext(filename)
            if f_extension.lower() not in ['.png', '.jpg', '.jpeg', '.gif', '.pgm']:
                continue

            path = os.path.join(subject_path, filename)
            img = cv2.imread(path, 0)
            if img is None:
                continue

            img = cv2.resize(img, (im_width, im_height))
            images.append(img)
            labels.append(id)
        id += 1

# Check if training data is available
if len(images) < 2:
    print("[ERROR] Not enough training data. At least 2 students are required.")
    exit()

# Convert lists to numpy arrays
(images, labels) = [np.array(lis) for lis in [images, labels]]

# Train the recognizer
model = cv2.face.LBPHFaceRecognizer_create()  # ✅ More robust than FisherFaces
model.train(images, labels)
print("[INFO] Training completed.")

# Load Haar Cascade from OpenCV path
fn_haar = cv2.data.haarcascades + "haarcascade_frontalface_default.xml"
haar_cascade = cv2.CascadeClassifier(fn_haar)

# Excel setup
if not os.path.exists(excel_file):
    wb = Workbook()
    ws = wb.active
    ws.append(["Name", "Date", "Time"])
    wb.save(excel_file)

def mark_attendance(name):
    now = datetime.now()
    date_str = now.strftime("%Y-%m-%d")
    time_str = now.strftime("%H:%M:%S")

    wb = load_workbook(excel_file)
    ws = wb.active
    ws.append([name, date_str, time_str])
    wb.save(excel_file)
    print(f"[INFO] Attendance marked for {name} at {time_str}")

# Start webcam
webcam = cv2.VideoCapture(0)
if not webcam.isOpened():
    print("[ERROR] Cannot access webcam.")
    exit()

while True:
    rval, frame = webcam.read()
    if not rval:
        print("[ERROR] Failed to grab frame.")
        break

    frame = cv2.flip(frame, 1, 0)
    gray = cv2.cvtColor(frame, cv2.COLOR_BGR2GRAY)
    mini = cv2.resize(gray, (int(gray.shape[1] / size), int(gray.shape[0] / size)))

    faces = haar_cascade.detectMultiScale(mini)
    for face_i in faces:
        (x, y, w, h) = [v * size for v in face_i]
        face = gray[y:y + h, x:x + w]
        face_resize = cv2.resize(face, (im_width, im_height))

        try:
            label_id, confidence = model.predict(face_resize)
            cv2.rectangle(frame, (x, y), (x + w, y + h), (0, 255, 0), 3)

            if confidence > confidence_threshold:
                cv2.putText(frame, "Unknown", (x, y - 10),
                            cv2.FONT_HERSHEY_PLAIN, 1.2, (0, 0, 255), 2)
            else:
                student_name = names[label_id]
                cv2.putText(frame, f'{student_name} ({confidence:.0f})',
                            (x, y - 10), cv2.FONT_HERSHEY_PLAIN, 1.2, (0, 255, 0), 2)

                # ✅ Save attendance in Excel
                mark_attendance(student_name)

        except Exception as e:
            print(f"[ERROR] Prediction failed: {e}")
            cv2.putText(frame, "Error", (x, y - 10),
                        cv2.FONT_HERSHEY_PLAIN, 1, (0, 0, 255), 2)

    cv2.imshow('Face Recognition Attendance', frame)
    key = cv2.waitKey(10)
    if key == 27:  # ESC key to break
        break

webcam.release()
cv2.destroyAllWindows()
