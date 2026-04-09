from openpyxl import load_workbook
import csv
from datetime import datetime

wb = load_workbook("attendance.xlsx")
ws = wb.active

today = datetime.now().strftime("%Y-%m-%d")

with open("attendance_today.csv", "w", newline="") as f:
    writer = csv.writer(f)
    writer.writerow(["S.No", "Student Name"])
    s_no = 1
    students = set()
    for row in ws.iter_rows(min_row=2, values_only=True):
        name, date, _ = row
        if date == today and name not in students:
            writer.writerow([s_no, name])
            students.add(name)
            s_no += 1
